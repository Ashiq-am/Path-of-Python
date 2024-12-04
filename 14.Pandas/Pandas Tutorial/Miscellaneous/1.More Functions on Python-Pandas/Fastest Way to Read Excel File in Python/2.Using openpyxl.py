from openpyxl import load_workbook

# Load workbook
wb = load_workbook('data.xlsx', read_only=True)

# Select a sheet
ws = wb['Sheet1']

# Read data
li = []
for i in ws.iter_rows(values_only=True):
    li.append(i)
print(li[:5])