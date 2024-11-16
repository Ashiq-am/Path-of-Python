from openpyxl import load_workbook

# Load the source Excel file
source_wb = load_workbook('excel1.xlsx')
source_ws = source_wb.active

# Create a new Excel file
destination_wb = load_workbook('excel1.xlsx')
destination_ws = destination_wb.active

# Copy data from source to destination
for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, values_only=True):
	destination_ws.append(row)

# Save the destination file
destination_wb.save('excel2.xlsx')
