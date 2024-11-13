# Importing openpyxl library
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule

# Create a workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Add some sample data
data = [
    [10, 20, 30, 40],
    [15, 25, 35, 45],
    [50, 60, 70, 80],
    [5, 10, 15, 20]
]

for row in data:
    ws.append(row)

# Apply conditional formatting based on cell values
# Example 1: Highlight cells greater than 40 in red
red_fill = PatternFill(start_color="FFEE1111",
                       end_color="FFEE1111", fill_type="solid")
ws.conditional_formatting.add('A1:D4', CellIsRule(
    operator='greaterThan', formula=['40'], stopIfTrue=True, fill=red_fill))

# Example 2: Apply a 2-color scale between 10 and 80
ws.conditional_formatting.add('A1:D4', ColorScaleRule(start_type='num', start_value=10, start_color='FFFFEDA0',
                                                      end_type='num', end_value=80, end_color='FF00BFFF'))

# Example 3: Formula-based conditional formatting (highlight even numbers in blue)
blue_fill = PatternFill(start_color="FF0000FF",
                        end_color="FF0000FF", fill_type="solid")
ws.conditional_formatting.add('A1:D4', FormulaRule(
    formula=['MOD(A1,2)=0'], fill=blue_fill))

# Save the workbook
wb.save('conditional_formatting_example.xlsx')
