import openpyxl
wb = openpyxl.load_workbook('merge.xlsx')
sheet = wb.active

# unmerge the cells
sheet.unmerge_cells('A2:D4')

sheet.unmerge_cells('C6:D6')

wb.save('merge.xlsx')
