# import modules
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
Spreadsheet = Workbook()

worksheet = Spreadsheet.active
wb = Workbook()
ws = wb.active

# assign title
ws['A1'] = 'Two logos demonstrated'

# create an image
img = Image('gfg.png')
img2=Image('gfg2.png')

# add image
ws.add_image(img, 'A1')
ws.add_image(img2, 'H1')
wb.save('logo.xlsx')
