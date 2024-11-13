# import modules
from openpyxl import Workbook
import datetime


wb = Workbook()

ws = wb.active

# assign sheet
ws['A1'] = 75

# add data
ws.append([1, 2, 3])
ws.append([4, 5, 6])
ws.append([8, 10, 12])

ws['A2'] = datetime.datetime.now()

# save spreadsheet
wb.save("test.xlsx")
