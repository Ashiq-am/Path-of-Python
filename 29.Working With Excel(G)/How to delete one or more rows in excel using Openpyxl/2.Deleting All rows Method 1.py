import openpyxl


def delete(sheet):

	# continuously delete row 2 untill there
	# is only a single row left over
	# that contains column names
	while(sheet.max_row > 1):
		# this method removes the row 2
		sheet.delete_rows(2)
	# return to main function
	return


if __name__ == '__main__':

		# enter your file path
	path = './delete_every_rows.xlsx'

	# load excel file
	book = openpyxl.load_workbook(path)

	# select the sheet
	sheet = book['sheet1']

	print("Maximum rows before removing:", sheet.max_row)

	delete(sheet)

	print("Maximum rows after removing:", sheet.max_row)

	# save the file to the path
	path = './openpy.xlsx'
	book.save(path)
