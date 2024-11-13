import openpyxl

if __name__ == '__main__':

	# enter your file path
	path = './delete_every_rows.xlsx'

	# load excel file
	book = openpyxl.load_workbook(path)

	# select the sheet
	sheet = book['sheet1']

	print("Maximum rows before removing:", sheet.max_row)

	# sheet.max_row is the maximum number
	# of rows that the sheet have
	# delete_row() method removes rows, first parameter represents row
	# number and sencond parameter represents number of rows
	# to delete from the row number
	sheet.delete_rows(2, sheet.max_row-1)

	print("Maximum rows after removing:", sheet.max_row)

	# save the file to the path
	path = './openpy.xlsx'
	book.save(path)
