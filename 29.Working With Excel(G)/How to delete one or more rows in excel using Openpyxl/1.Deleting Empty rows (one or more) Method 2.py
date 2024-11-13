# import openpyxl library
import openpyxl


# function to remove empty rows

def remove(sheet):
    # iterate the sheet by rows
    for row in sheet.iter_rows():

        # all() return False if all of the row value is None
        if not all(cell.value for cell in row):
            # detele the empty row
            sheet.delete_rows(row[0].row, 1)

            remove(sheet)


            return

if __name__ == '__main__':

    # enter your file path
    path = './delete_empty_rows.xlsx'

    # load excel file
    book = openpyxl.load_workbook(path)

    # select the sheet
    sheet = book['daily sales']

    print("Maximum rows before removing:", sheet.max_row)

    # iterate the sheet
    for row in sheet:
        remove(sheet)

    print("Maximum rows after removing:", sheet.max_row)

    # save the file to the path
    path = './openpy.xlsx'
    book.save(path)
