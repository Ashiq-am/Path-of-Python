from openpyxl import load_workbook
from json import dumps

# Load Excel workbook
wb = load_workbook("student.xlsx")

# Choose a specific sheet
sheet = wb["suraj1"]

# Find the number of rows and columns in the sheet
rows = sheet.max_row
columns = sheet.max_column

# List to store all rows as dictionaries
lst = []

# Iterate over rows and columns to extract data
for i in range(1, rows):
	row = {}
	for j in range(1, columns):
		column_name = sheet.cell(row=1, column=j)
		row_data = sheet.cell(row=i+1, column=j)

		row.update(
			{
				column_name.value: row_data.value
			}
		)
	lst.append(row)

# Convert extracted data into JSON format
json_data = dumps(lst)

# Print the JSON data
print(json_data)
