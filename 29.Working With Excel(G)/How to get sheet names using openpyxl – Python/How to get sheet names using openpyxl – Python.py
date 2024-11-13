# importing the openpyxl module as xl
import openpyxl as xl

# excel file used here gfg.xlsx
excel_file = "Example.xlsx"

# load the workbook
wb = xl.load_workbook(excel_file)

# print the list that stores the sheetnames
print(wb.sheetnames)

# we can also get the name of the sheet which is active by using the active property
print("Active sheet: ", wb.active)

# By default the active sheet is the 1st one and can be changed
wb._active_sheet_index = 4
sheet = wb.active
print("Active sheet: ", wb.active)
