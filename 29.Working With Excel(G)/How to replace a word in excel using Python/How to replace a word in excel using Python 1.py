# Reading and writing in excel can be done by single module
import openpyxl
from openpyxl.utils.cell import get_column_letter

workbook = openpyxl.load_workbook('sampleexcelopenpyxl.xlsx')
workbook.sheetnames
worksheet = workbook["Test"]

# Number of rows
number_of_rows = worksheet.max_row

# Number of columns
number_of_columns = worksheet.max_column

replacementTextKeyPairs = {'1': 'One', '2': 'Two', '3': 'Three'}

# Iterate over the columns and rows, search
# for the text and replace
for i in range(number_of_columns):
    for k in range(number_of_rows):

        cellValue = str(worksheet[get_column_letter(i + 1) + str(k + 1)].value)

        for key in replacementTextKeyPairs.keys():

            if str(cellValue) == key:
                newCellValue = replacementTextKeyPairs.get(key)
                worksheet[get_column_letter(i + 1) + str(k + 1)] = str(newCellValue)

workbook.save('sampleexcelwithreplacedtextusingopenpyxl.xlsx')
