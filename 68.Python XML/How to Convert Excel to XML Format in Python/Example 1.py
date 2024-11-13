# Install the openpyxl library
from openpyxl import load_workbook

# Loading our Excel file
wb = load_workbook("demo_database.xlsx")

# creating the sheet 1 object
ws = wb.worksheets[0]

# Iterating rows for getting the values of each row
for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
	print([cell.value for cell in row])
